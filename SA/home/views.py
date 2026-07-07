"""
App 'home' - đăng nhập / phân quyền / trang chủ dùng chung.

============================================================
LUỒNG TRUY CẬP (đã chốt)
============================================================
1. Chưa đăng nhập web mẹ (không có session)
   -> parent_username = None
   -> no_login.html (nút sang web mẹ đăng nhập)

2. Đã đăng nhập web mẹ, NHƯNG username không tồn tại trong auth_user (db con)
   -> no_permission.html ("đã đăng nhập nhưng chưa được cấp quyền")

3. Đã đăng nhập web mẹ + username có trong auth_user (db con)
   -> dashboard.html

============================================================
ĐĂNG XUẤT (single logout - thoát cả hệ thống)
============================================================
Web con KHÔNG tự xóa được cookie session của web mẹ. Nên nút Đăng xuất
sẽ: (1) xóa phiên web con, rồi (2) chuyển hướng sang URL đăng xuất của
web mẹ để web mẹ tự xóa phiên của nó -> user thoát khỏi toàn hệ thống.

Phụ thuộc:
- request.parent_username do ParentSessionMiddleware (package gốc) gán.
"""
from django.shortcuts import render, redirect
from django.contrib.auth import logout as django_logout
from django.conf import settings

from .services import get_username_from_request, user_exists


# === Cấu hình (nên đặt trong settings.py, không sửa file này) ===
SCM_LOGIN_URL = getattr(settings, 'SCM_LOGIN_URL', 'http://10.92.184.241:8888/accounts/login/')
SCM_LOGOUT_URL = getattr(settings, 'SCM_LOGOUT_URL', 'http://10.92.184.241:8888/accounts/logout/')
PROJECT_NAME = getattr(settings, 'PROJECT_NAME', 'Shipping Advice')


def dashboard(request):
    """Trang chủ - kiểm tra đăng nhập (web mẹ) rồi kiểm tra user tồn tại (db con)."""
    username = get_username_from_request(request)

    # 1. Chưa đăng nhập web mẹ
    if not username:
        return render(request, 'home/no_login.html', {
            'scm_login_url': SCM_LOGIN_URL,
            'project_name': PROJECT_NAME,
        })

    # 2. Đã đăng nhập web mẹ, nhưng username không có trong auth_user (db con)
    if not user_exists(username):
        return render(request, 'home/no_permission.html', {
            'username': username,
            'project_name': PROJECT_NAME,
        })

    # 3. Có trong db con -> dashboard
    return render(request, 'home/dashboard.html', {
        'username': username,
        'project_name': PROJECT_NAME,
    })


def logout(request):
    """
    Đăng xuất cả web con lẫn web mẹ, nhưng KHÔNG chuyển trang sang web mẹ.

    Bước 1: django_logout() xóa phiên web con (cookie SA_sessionid).
    Bước 2: render trang trung gian logout.html. Trang này có thẻ <img> ẩn
            trỏ tới URL logout web mẹ -> trình duyệt tự tải ngầm (kèm cookie
            web mẹ vì cùng host) -> web mẹ xóa phiên của nó.
    Bước 3: trang tự chuyển về trang chủ web con (lúc này cả 2 phiên đã hết
            -> hiện "Chưa đăng nhập").

    User luôn ở lại web con, không bị nhảy sang giao diện web mẹ.
    """
    django_logout(request)
    return render(request, 'home/logout.html', {
        'scm_logout_url': SCM_LOGOUT_URL,
        'project_name': PROJECT_NAME,
    })

def set_language(request):
    """
    Đổi ngôn ngữ giao diện. Nhận ?lang=vi|en|ja, lưu vào session (nhớ cho
    lần sau), rồi quay về trang vừa xem.
    """
    from .translations import LANG_CODES

    lang = request.GET.get('lang')
    if lang in LANG_CODES:
        request.session['lang'] = lang
    back = request.META.get('HTTP_REFERER') or '/'
    return redirect(back)



# =========================================================================
# DEMO TÌM KIẾM (2 kiểu) - tham khảo, có thể xóa khi không cần
# =========================================================================

def demo_search_client(request):
    """Demo kiểu 1: lọc text trên màn hình. View chỉ render trang tĩnh;
    việc lọc do JavaScript ở base.html làm (chế độ 'client')."""
    return render(request, 'home/demo_search_client.html', {
        'project_name': PROJECT_NAME,
        'username': get_username_from_request(request),
    })


def demo_search_server(request):
    """Demo kiểu 2: tìm trong database. Nhận ?q= rồi truy vấn.
    Dùng bảng auth_user làm dữ liệu mẫu (có sẵn, khỏi tạo model)."""
    from django.contrib.auth import get_user_model
    from django.db.models import Q

    User = get_user_model()
    query = (request.GET.get('q') or '').strip()

    results = []
    if query:
        # icontains = tìm gần đúng, không phân biệt hoa/thường.
        # Q(...) | Q(...) = khớp MỘT TRONG các trường.
        results = User.objects.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        ).order_by('username')[:50]     # giới hạn 50 để tránh trả quá nhiều

    return render(request, 'home/demo_search_server.html', {
        'project_name': PROJECT_NAME,
        'username': get_username_from_request(request),
        'query': query,
        'results': results,
    })


# =========================================================================
# LIVE SEARCH (gợi ý xổ xuống + mở link tab mới)
# =========================================================================

def demo_live_search(request):
    """Trang demo live search - chỉ render giao diện. Việc gợi ý do JS gọi
    tới endpoint live_search_api bên dưới."""
    return render(request, 'home/demo_live_search.html', {
        'project_name': PROJECT_NAME,
        'username': get_username_from_request(request),
    })


def live_search_api(request):
    """
    Endpoint trả JSON cho live search. Nhận ?q=... , trả danh sách gợi ý.

    Mỗi gợi ý có: tiêu đề, mô tả phụ, và URL để mở (tab mới).
    Ở đây dùng DỮ LIỆU MẪU cứng cho dễ hình dung. Trong dự án thật, thay
    phần SAMPLE bằng truy vấn database của bạn và ghép URL tương ứng.
    """
    from django.http import JsonResponse

    query = (request.GET.get('q') or '').strip().lower()

    # === DỮ LIỆU MẪU - thay bằng truy vấn database thật ===
    SAMPLE = [
        {'title': 'Exchange Rate into SAP', 'tag': 'FICO', 'code': '0B08',
         'desc': r'\\10.92.184.241\share\Auto\FICO\0B08\Exchange_rate',
         'url': 'http://10.92.184.241:8019/demo/search-client/'},
        {'title': 'Material Master Upload', 'tag': 'MM', 'code': 'MM01',
         'desc': r'\\10.92.184.241\share\Auto\MM\MM01\Material',
         'url': 'http://10.92.184.241:8019/demo/search-server/'},
        {'title': 'Vendor Master Create', 'tag': 'MM', 'code': 'XK01',
         'desc': r'\\10.92.184.241\share\Auto\MM\XK01\Vendor',
         'url': 'http://10.92.184.241:8019/'},
        {'title': 'Post FI Document', 'tag': 'FICO', 'code': 'FB01',
         'desc': r'\\10.92.184.241\share\Auto\FICO\FB01\Posting',
         'url': 'http://10.92.184.241:8019/demo/search-server/'},
    ]

    if not query:
        results = []
    else:
        # Lọc gần đúng: khớp nếu query nằm trong tiêu đề, tag, hoặc mã.
        results = [
            item for item in SAMPLE
            if query in item['title'].lower()
            or query in item['tag'].lower()
            or query in item['code'].lower()
        ]

    return JsonResponse({'results': results})




# =========================================================================
# GRIDVIEW - 3 dạng bảng dữ liệu (tham khảo, có thể xóa)
# Dữ liệu mẫu dùng chung; dự án thật thay bằng truy vấn database.
# =========================================================================

def _sample_rows():
    """Dữ liệu mẫu cho các bảng demo (thay bằng Model.objects... khi dùng thật).

    Sinh tự động >50 dòng bằng vòng lặp để thử phân trang. Khi dùng thật,
    thay toàn bộ hàm này bằng truy vấn database, vd:
        return list(Route.objects.values('id','code','name','vendor','days','status'))
    """
    cities = ['Hà Nội', 'Hải Phòng', 'Đà Nẵng', 'HCM', 'Cần Thơ', 'Huế',
              'Vũng Tàu', 'Đà Lạt', 'Lào Cai', 'Nha Trang', 'Quy Nhơn',
              'Buôn Ma Thuột', 'Vinh', 'Hạ Long', 'Phú Quốc']
    vendors = ['Vendor A', 'Vendor B', 'Vendor C', 'Vendor D', 'Vendor E']
    statuses = ['ok', 'ok', 'ok', 'warn', 'danger']   # 'ok' nhiều hơn cho thực tế

    rows = []
    for i in range(1, 610000):        # 60 dòng
        origin = cities[i % len(cities)]
        dest = cities[(i * 3 + 2) % len(cities)]
        rows.append({
            'id': i,
            'code': 'RT%03d' % i,             # RT001, RT002, ...
            'name': '%s - %s' % (origin, dest),
            'vendor': vendors[i % len(vendors)],
            'days': (i % 5) + 1,              # 1..5
            'status': statuses[i % len(statuses)],
        })
    return rows


def demo_grid_display(request):
    """Dạng 1: bảng hiển thị + sắp xếp + phân trang (bằng Django/server)."""
    from django.core.paginator import Paginator

    rows = _sample_rows()

    # --- Tìm kiếm (server): lọc TOÀN BỘ dữ liệu theo ?q= ---
    # Khi dùng thật, nên lọc ở database: Route.objects.filter(Q(code__icontains=q)|...)
    # thay vì lọc list trong Python (nhanh hơn với dữ liệu lớn).
    q = (request.GET.get('gq') or '').strip().lower()
    if q:
        # Tìm TẤT CẢ cột. Muốn giới hạn -> thêm fields, vd:
        #   _match_row(r, q, fields=['code', 'name', 'vendor'])
        rows = [r for r in rows if _match_row(r, q)]

    # --- Sắp xếp. Mặc định cho sort HẾT cột.
    #     Muốn giới hạn -> _sort_rows(rows, request, sortable=['code','days'])
    #     và truyền 'sortable_cols' cùng danh sách đó ra template. ---
    rows, sort, direction = _sort_rows(rows, request)
    SORTABLE_COLS = None   # None = sort hết; hoặc ['code','name'] để giới hạn

    # --- Phân trang ---
    PER_PAGE = _get_per_page(request)
    paginator = Paginator(rows, PER_PAGE)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'home/demo_grid_display.html', {
        'username': get_username_from_request(request),
        'page': page,
        'page_window': _page_window(page),
        'sort': sort,
        'dir': direction,
        'q': q,
        'per_page': PER_PAGE,
        'per_page_choices': PER_PAGE_CHOICES,
        'sortable_cols': SORTABLE_COLS,
    })


PER_PAGE_CHOICES = [25, 50, 100, 200, 500]
DEFAULT_PER_PAGE = 100


def _get_per_page(request):
    """Đọc số dòng/trang do người dùng chọn (?per_page=), chỉ nhận giá trị
    trong PER_PAGE_CHOICES; ngoài danh sách -> dùng mặc định."""
    try:
        val = int(request.GET.get('per_page', DEFAULT_PER_PAGE))
    except (TypeError, ValueError):
        val = DEFAULT_PER_PAGE
    return val if val in PER_PAGE_CHOICES else DEFAULT_PER_PAGE


def _page_window(page, width=2):
    """
    Trả về danh sách số trang cần hiện: vài trang quanh trang hiện tại,
    cộng trang đầu/cuối, chèn None ở chỗ bị lược (hiển thị thành '...').
    Ví dụ trang 25/50 -> [1, None, 23, 24, 25, 26, 27, None, 50].
    width = số trang hiện mỗi bên quanh trang hiện tại.
    """
    current = page.number
    total = page.paginator.num_pages

    # Tập các trang cần hiện: đầu, cuối, và quanh trang hiện tại
    pages = set([1, total])
    for p in range(current - width, current + width + 1):
        if 1 <= p <= total:
            pages.add(p)

    # Sắp xếp và chèn None vào chỗ có khoảng trống (>1 bậc)
    result = []
    prev = 0
    for p in sorted(pages):
        if prev and p - prev > 1:
            result.append(None)      # None -> render thành '...'
        result.append(p)
        prev = p
    return result


# Tất cả cột có thể sort (khi trang muốn "sort hết")
_ALL_SORT_COLS = {'code', 'name', 'vendor', 'days', 'status'}


def _sort_rows(rows, request, sortable=None):
    """
    Sắp xếp rows theo ?sort=&dir=. Hỗ trợ 2 cách giống tìm kiếm:

    CÁCH 1 - Sort HẾT các cột (mặc định, sortable=None):
        _sort_rows(rows, request)

    CÁCH 2 - Chỉ CHO SORT vài cột:
        _sort_rows(rows, request, sortable=['code', 'days'])
        -> bấm cột ngoài danh sách sẽ bị bỏ qua (không sort).

    Trả về (rows đã sắp xếp, sort, direction).
    """
    allowed = _ALL_SORT_COLS if sortable is None else set(sortable)
    sort = request.GET.get('sort', '')
    direction = request.GET.get('dir', 'asc')
    if sort in allowed:
        rows = sorted(rows, key=lambda r: r[sort], reverse=(direction == 'desc'))
    else:
        sort = ''   # cột không được phép -> coi như chưa sort
    return rows, sort, direction


def _match_row(row, q, fields=None):
    """
    Kiểm tra 1 dòng có khớp từ khóa q không. Hỗ trợ 2 cách:

    CÁCH 1 - Tìm TẤT CẢ cột (mặc định, fields=None):
        _match_row(row, q)
        -> ghép mọi giá trị (trừ 'id') thành chuỗi rồi tìm.

    CÁCH 2 - Tìm CHỈ các cột chỉ định (truyền fields):
        _match_row(row, q, fields=['code', 'name', 'vendor'])
        -> chỉ tìm trong các cột liệt kê.

    Đổi cách nào tùy trang: gọi view truyền fields hoặc không.
    """
    if fields is None:
        # CÁCH 1: quét tất cả cột (trừ id)
        haystack = ' '.join(str(v).lower() for k, v in row.items() if k != 'id')
    else:
        # CÁCH 2: chỉ các cột chỉ định
        haystack = ' '.join(str(row.get(f, '')).lower() for f in fields)
    return q in haystack


def _filter_sort_rows(request, search_fields=None):
    """Lọc + sắp xếp dữ liệu theo tham số hiện tại (dùng chung cho grid + export).
    search_fields=None -> tìm tất cả cột; hoặc truyền list cột để giới hạn.
    Trả về list dòng đã lọc/sắp xếp (CHƯA phân trang)."""
    rows = _sample_rows()
    q = (request.GET.get('gq') or '').strip().lower()
    if q:
        rows = [r for r in rows if _match_row(r, q, search_fields)]
    sort = request.GET.get('sort', '')
    direction = request.GET.get('dir', 'asc')
    if sort in {'code', 'name', 'vendor', 'days', 'status'}:
        rows = sorted(rows, key=lambda r: r[sort], reverse=(direction == 'desc'))
    return rows


# Nhãn cột khi xuất file (thứ tự cột + tiêu đề)
_EXPORT_COLUMNS = [
    ('code', 'Mã Route'),
    ('name', 'Tuyến'),
    ('vendor', 'Vendor'),
    ('days', 'Số ngày'),
    ('status', 'Trạng thái'),
]


def grid_export(request):
    """Xuất dữ liệu grid theo bộ lọc/tìm hiện tại. ?format=csv|xlsx.
    (PDF chưa làm - cần thư viện ngoài.)"""
    fmt = request.GET.get('format', 'csv')
    rows = _filter_sort_rows(request)

    if fmt == 'xlsx':
        return _export_xlsx(rows)
    return _export_csv(rows)   # mặc định csv


def _export_csv(rows):
    """Xuất CSV. Dùng thư viện csv có sẵn của Python (không cần cài gì).
    BOM utf-8-sig để Excel mở tiếng Việt không lỗi font."""
    import csv
    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="routes.csv"'
    response.write('\ufeff')   # BOM cho Excel đọc UTF-8

    writer = csv.writer(response)
    writer.writerow([label for _, label in _EXPORT_COLUMNS])   # tiêu đề
    for r in rows:
        writer.writerow([r[key] for key, _ in _EXPORT_COLUMNS])
    return response


def _export_xlsx(rows):
    """Xuất Excel .xlsx bằng openpyxl (đã có trong requirements)."""
    from django.http import HttpResponse
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        # openpyxl chưa cài -> báo nhẹ, không vỡ
        from django.http import HttpResponse as _R
        return _R('Cần cài openpyxl: pip install openpyxl', status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Routes'

    # Tiêu đề in đậm, nền màu
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0A4A8F')
    for col_idx, (_, label) in enumerate(_EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill

    # Dữ liệu
    for row_idx, r in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(_EXPORT_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=r[key])

    # Giãn cột cho dễ đọc
    for col_idx, (key, label) in enumerate(_EXPORT_COLUMNS, start=1):
        width = max(len(str(label)), 14)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="routes.xlsx"'
    wb.save(response)
    return response


def demo_grid_crud(request):
    """Dạng 2: bảng có nút Thêm/Sửa/Xóa (CRUD) + sort + phân trang.
    Thao tác Sửa/Xóa qua trang riêng nên không xung đột phân trang."""
    from django.core.paginator import Paginator

    rows = _sample_rows()

    # Tìm kiếm (server) theo ?q=
    q = (request.GET.get('gq') or '').strip().lower()
    if q:
        # Tìm TẤT CẢ cột. Muốn giới hạn -> _match_row(r, q, fields=[...])
        rows = [r for r in rows if _match_row(r, q)]

    # Sắp xếp - VÍ DỤ Cách 2: chỉ cho sort vài cột (code, vendor, days).
    # Cột 'name' và 'status' sẽ hiện tiêu đề tĩnh, không bấm sort được.
    SORTABLE_COLS = ['code', 'vendor', 'days']
    rows, sort, direction = _sort_rows(rows, request, sortable=SORTABLE_COLS)

    # Phân trang
    PER_PAGE = _get_per_page(request)
    paginator = Paginator(rows, PER_PAGE)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'home/demo_grid_crud.html', {
        'username': get_username_from_request(request),
        'page': page,
        'page_window': _page_window(page),
        'sort': sort,
        'dir': direction,
        'q': q,
        'per_page': PER_PAGE,
        'per_page_choices': PER_PAGE_CHOICES,
        'sortable_cols': SORTABLE_COLS,
    })


def demo_grid_editable(request):
    """Dạng 3: bảng chỉnh sửa trực tiếp (editable). Sửa tại ô, bấm Lưu ->
    gửi toàn bộ thay đổi 1 lần. Demo hiển thị giao diện + JS thu thập thay đổi."""
    return render(request, 'home/demo_grid_editable.html', {
        'username': get_username_from_request(request),
        'rows': _sample_rows(),
    })


def route_form(request, pk=None):
    """
    Form Thêm/Sửa dùng chung. pk=None -> Thêm mới; pk có giá trị -> Sửa.

    Luồng chuẩn của Django Form:
    - GET  : hiện form (rỗng khi thêm, điền sẵn dữ liệu khi sửa).
    - POST : nhận dữ liệu -> form.is_valid() tự validate:
             + hợp lệ  -> lưu (ở đây demo chỉ in ra) rồi redirect.
             + lỗi     -> render lại form kèm thông báo lỗi từng trường.
    """
    from .forms import RouteForm

    is_edit = pk is not None

    # Khi SỬA: nạp dữ liệu cũ vào form (demo dùng dữ liệu mẫu).
    # Thật: instance = get_object_or_404(Route, pk=pk); form = RouteForm(instance=...)
    initial = {}
    if is_edit:
        for r in _sample_rows():
            if str(r['id']) == str(pk):
                initial = {
                    'code': r['code'], 'name': r['name'],
                    'days': r['days'], 'vendor': r.get('vendor', ''),
                    'status': r['status'],
                }
                break

    if request.method == 'POST':
        form = RouteForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            # DEMO: chỉ in ra. Thật: tạo/cập nhật bản ghi rồi lưu database.
            print('Dữ liệu hợp lệ, sẽ lưu:', data)
            # Sau khi lưu -> quay về danh sách
            return redirect('home:demo_grid_crud')
        # Không hợp lệ -> rơi xuống, render lại form (đã có lỗi đính kèm)
    else:
        form = RouteForm(initial=initial)

    return render(request, 'home/route_form.html', {
        'username': get_username_from_request(request),
        'form': form,
        'is_edit': is_edit,
        'back_url': '/demo/grid-crud/',
    })
